import openpyxl
import os
from utils.crypto import encriptar, desencriptar


def carregar_excel(nome_bd, fernet):
    if os.path.exists(nome_bd):
        desencriptar(nome_bd, fernet)
        wb = openpyxl.load_workbook(nome_bd)
        ws = wb.active
        encriptar(nome_bd, fernet)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Passwords"
        ws.append(["Descrição", "Password"])

    return wb, ws


def get_passwords(nome_bd, fernet):
    desencriptar(nome_bd, fernet)
    wb = openpyxl.load_workbook(nome_bd)
    ws = wb.active
    encriptar(nome_bd, fernet)
    return list(ws.iter_rows(min_row=2, values_only=True))


def apagar_password(nome_bd, fernet, indice):
    desencriptar(nome_bd, fernet)
    wb = openpyxl.load_workbook(nome_bd)
    ws = wb.active
    ws.delete_rows(indice + 1)
    wb.save(nome_bd)
    encriptar(nome_bd, fernet)


def editar_password(nome_bd, fernet, indice, nova_descricao, nova_password):
    desencriptar(nome_bd, fernet)
    wb = openpyxl.load_workbook(nome_bd)
    ws = wb.active
    ws.cell(row=indice + 1, column=1).value = nova_descricao
    ws.cell(row=indice + 1, column=2).value = nova_password
    wb.save(nome_bd)
    encriptar(nome_bd, fernet)


def adicionar_password(nome_bd, fernet, descricao, password):
    desencriptar(nome_bd, fernet)
    wb = openpyxl.load_workbook(nome_bd)
    ws = wb.active
    ws.append([descricao, password])
    wb.save(nome_bd)
    encriptar(nome_bd, fernet)
